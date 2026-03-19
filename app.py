from flask import Flask, jsonify, request, render_template, session, redirect, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import urllib.request
import urllib.parse
import json as _json
from datetime import datetime, timezone, timedelta
from functools import wraps
import bcrypt
import re
import os
import threading
import uuid
import mimetypes
import io
from werkzeug.utils import secure_filename

# Fuso horário Brasil (UTC-3)
BR_TZ = timezone(timedelta(hours=-3))

def agora_br():
    return datetime.now(BR_TZ).replace(tzinfo=None)

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ─────────────────────────────────────────
# CONFIGURACAO
# ─────────────────────────────────────────
ADMIN_MASTER_EMAIL = 'henriquecipriani@gmail.com'

def get_database_url():
    url = os.environ.get('DATABASE_URL', 'sqlite:///taskflow.db')
    if url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)
    return url

app = Flask(__name__)
app.config['SECRET_KEY']                     = os.environ.get('SECRET_KEY', 'taskflow-dev-secret-troque-em-producao')
app.config['SQLALCHEMY_DATABASE_URI']        = get_database_url()
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

if os.environ.get('FLASK_ENV') == 'production':
    app.config['SESSION_COOKIE_SECURE']   = True
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

SUPABASE_URL    = os.environ.get('SUPABASE_URL', '')
SUPABASE_KEY    = os.environ.get('SUPABASE_SERVICE_KEY', '')
SUPABASE_BUCKET = os.environ.get('SUPABASE_BUCKET', 'anexos-taskflow')
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024

APP_URL = os.environ.get('APP_URL', 'https://projeto-zvam.onrender.com')

EXTENSOES_PERMITIDAS = {
    'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx',
    'txt', 'csv', 'zip', 'rar', '7z',
    'png', 'jpg', 'jpeg', 'gif', 'webp',
    'mp4', 'mov', 'avi'
}

STATUSES_VALIDOS    = ['Não iniciado', 'Iniciado', 'Em andamento', 'Pausado', 'Aguardo retorno', 'Finalizado']
PRIORIDADES_VALIDAS = ['Nenhuma', 'Alta']

db   = SQLAlchemy(app)
CORS(app)

# ─────────────────────────────────────────
# TABELAS ASSOCIATIVAS
# ─────────────────────────────────────────
tarefa_responsaveis = db.Table('tarefa_responsaveis',
    db.Column('tarefa_codigo', db.Integer, db.ForeignKey('tarefas.codigo'), primary_key=True),
    db.Column('usuario_id',    db.Integer, db.ForeignKey('usuarios.id'),    primary_key=True)
)

tarefa_admins = db.Table('tarefa_admins',
    db.Column('tarefa_codigo', db.Integer, db.ForeignKey('tarefas.codigo'), primary_key=True),
    db.Column('usuario_id',    db.Integer, db.ForeignKey('usuarios.id'),    primary_key=True)
)

from flask_migrate import Migrate
migrate = Migrate(app, db)

# ─────────────────────────────────────────
# VALIDAÇÃO
# ─────────────────────────────────────────
def validar_senha_forte(senha):
    if len(senha) < 8:
        return False, 'A senha deve ter pelo menos 8 caracteres'
    if not re.search(r'[A-Z]', senha):
        return False, 'A senha deve conter pelo menos 1 letra maiuscula'
    if not re.search(r'[a-z]', senha):
        return False, 'A senha deve conter pelo menos 1 letra minuscula'
    if not re.search(r'\d', senha):
        return False, 'A senha deve conter pelo menos 1 numero'
    if not re.search(r'[!@#$%^&*(),.?\":{}|<>\-_=+\[\]\\;\'`~/]', senha):
        return False, 'A senha deve conter pelo menos 1 caractere especial (!@#$%^&* etc.)'
    return True, None

# ─────────────────────────────────────────
# MODELOS
# ─────────────────────────────────────────
class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id            = db.Column(db.Integer, primary_key=True)
    nome          = db.Column(db.String(100), nullable=False)
    funcao        = db.Column(db.String(100), nullable=False)
    email         = db.Column(db.String(150), unique=True, nullable=False)
    senha_hash    = db.Column(db.String(255), nullable=False)
    tipo_perfil   = db.Column(db.String(20), nullable=False)
    trocar_senha  = db.Column(db.Boolean, default=False)
    empresa       = db.Column(db.String(150), nullable=True)
    setor         = db.Column(db.String(150), nullable=True)
    comentarios   = db.relationship('Comentario', backref='autor', lazy=True)
    last_seen     = db.Column(db.DateTime, nullable=True)  # online indicator

    def definir_senha(self, senha_plain):
        self.senha_hash = bcrypt.hashpw(
            senha_plain.encode('utf-8'), bcrypt.gensalt()
        ).decode('utf-8')

    def verificar_senha(self, senha_plain):
        return bcrypt.checkpw(
            senha_plain.encode('utf-8'), self.senha_hash.encode('utf-8')
        )

    def is_admin(self):
        return self.tipo_perfil in ('Administrador', 'Admin Master')

    def is_master(self):
        return self.tipo_perfil == 'Admin Master'

    def to_dict(self):
        return {
            'id':           self.id,
            'nome':         self.nome,
            'funcao':       self.funcao,
            'email':        self.email,
            'tipo_perfil':  self.tipo_perfil,
            'trocar_senha': self.trocar_senha,
            'empresa':      self.empresa or '',
            'setor':        self.setor or '',
            'last_seen':    self.last_seen.isoformat() if self.last_seen else None
        }


class Tarefa(db.Model):
    __tablename__ = 'tarefas'
    codigo         = db.Column(db.Integer, primary_key=True, autoincrement=True)
    descricao      = db.Column(db.Text, nullable=False)
    data_criacao   = db.Column(db.DateTime, default=agora_br)
    status         = db.Column(db.String(30), default='Não iniciado')
    prioridade     = db.Column(db.String(10), default='Nenhuma')
    compartilhada  = db.Column(db.Boolean, default=True)
    criado_por     = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)
    empresa        = db.Column(db.String(150), nullable=True)
    deletado_em    = db.Column(db.DateTime, nullable=True)   # soft delete
    deletado_por   = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)
    data_prazo     = db.Column(db.Date, nullable=True)        # prazo
    recorrente     = db.Column(db.String(10), nullable=True)  # 'semanal' | 'mensal' | None
    responsaveis   = db.relationship('Usuario', secondary=tarefa_responsaveis, lazy='subquery',
                                     backref=db.backref('tarefas_responsavel', lazy=True),
                                     foreign_keys=[tarefa_responsaveis.c.tarefa_codigo,
                                                   tarefa_responsaveis.c.usuario_id])
    admins_colabs  = db.relationship('Usuario', secondary=tarefa_admins, lazy='subquery',
                                     backref=db.backref('tarefas_admin_colab', lazy=True),
                                     foreign_keys=[tarefa_admins.c.tarefa_codigo,
                                                   tarefa_admins.c.usuario_id])
    comentarios    = db.relationship('Comentario', backref='tarefa', lazy=True, cascade='all, delete-orphan')
    anexos         = db.relationship('Anexo', backref='tarefa', lazy=True, cascade='all, delete-orphan',
                                     foreign_keys='Anexo.id_tarefa')
    checklist      = db.relationship('ChecklistItem', backref='tarefa', lazy=True, cascade='all, delete-orphan',
                                     foreign_keys='ChecklistItem.id_tarefa',
                                     order_by='ChecklistItem.ordem')

    def to_dict(self, viewer_id=None):
        delegada = False
        comigo   = False
        if viewer_id is not None:
            resp_ids        = [u.id for u in self.responsaveis]
            admin_colab_ids = [u.id for u in self.admins_colabs]
            if self.criado_por == viewer_id and self.compartilhada and resp_ids:
                delegada = True
            if viewer_id in resp_ids or viewer_id in admin_colab_ids:
                comigo = True
        hoje = agora_br().date()
        prazo_status = None
        if self.data_prazo and self.status != 'Finalizado':
            diff = (self.data_prazo - hoje).days
            if diff < 0:      prazo_status = 'vencida'
            elif diff == 0:   prazo_status = 'hoje'
            elif diff <= 3:   prazo_status = 'urgente'
            else:             prazo_status = 'ok'
        return {
            'codigo':        self.codigo,
            'descricao':     self.descricao,
            'data_criacao':  self.data_criacao.strftime('%d/%m/%Y %H:%M'),
            'status':        self.status,
            'prioridade':    self.prioridade,
            'compartilhada': self.compartilhada,
            'criado_por':    self.criado_por,
            'empresa':       self.empresa or '',
            'delegada':      delegada,
            'comigo':        comigo,
            'deletado_em':   self.deletado_em.strftime('%d/%m/%Y %H:%M') if self.deletado_em else None,
            'deletado_por':  self.deletado_por,
            'data_prazo':    self.data_prazo.strftime('%Y-%m-%d') if self.data_prazo else None,
            'data_prazo_fmt':self.data_prazo.strftime('%d/%m/%Y') if self.data_prazo else None,
            'prazo_status':  prazo_status,
            'recorrente':    self.recorrente or None,
            'responsaveis':  [
                {'id': u.id, 'nome': u.nome, 'funcao': u.funcao, 'setor': u.setor or ''}
                for u in self.responsaveis
            ],
            'admins_colabs': [
                {'id': u.id, 'nome': u.nome, 'funcao': u.funcao}
                for u in self.admins_colabs
            ],
            'anexos_count':           len(self.anexos),
            'checklist_total':        len(self.checklist),
            'checklist_concluidos':   sum(1 for i in self.checklist if i.concluido)
        }


class Comentario(db.Model):
    __tablename__ = 'comentarios'
    id         = db.Column(db.Integer, primary_key=True)
    id_tarefa  = db.Column(db.Integer, db.ForeignKey('tarefas.codigo'), nullable=False)
    id_usuario = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    texto      = db.Column(db.Text, nullable=False)
    tipo       = db.Column(db.String(20), default='comentario')
    data_hora  = db.Column(db.DateTime, default=agora_br)

    def to_dict(self):
        return {
            'id':           self.id,
            'texto':        self.texto,
            'tipo':         self.tipo,
            'data_hora':    self.data_hora.strftime('%d/%m/%Y %H:%M:%S'),
            'autor_nome':   self.autor.nome if self.autor else 'Sistema',
            'autor_perfil': self.autor.tipo_perfil if self.autor else 'sistema'
        }


class Anexo(db.Model):
    __tablename__ = 'anexos'
    id            = db.Column(db.Integer, primary_key=True)
    id_tarefa     = db.Column(db.Integer, db.ForeignKey('tarefas.codigo'), nullable=False)
    id_usuario    = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    nome_original = db.Column(db.String(255), nullable=False)
    nome_arquivo  = db.Column(db.String(255), nullable=False)
    tamanho       = db.Column(db.Integer, nullable=False)
    mime_type     = db.Column(db.String(100), nullable=True)
    data_upload   = db.Column(db.DateTime, default=agora_br)
    uploader      = db.relationship('Usuario', foreign_keys=[id_usuario])

    def to_dict(self):
        url = self.nome_arquivo if self.nome_arquivo.startswith('http') else f'/api/anexos/{self.id}/download'
        return {
            'id':            self.id,
            'nome_original': self.nome_original,
            'tamanho':       self.tamanho,
            'mime_type':     self.mime_type or '',
            'data_upload':   self.data_upload.strftime('%d/%m/%Y %H:%M'),
            'uploader_nome': self.uploader.nome if self.uploader else 'Desconhecido',
            'url_download':  url
        }


class ChecklistItem(db.Model):
    __tablename__ = 'checklist_items'
    id            = db.Column(db.Integer, primary_key=True)
    id_tarefa     = db.Column(db.Integer, db.ForeignKey('tarefas.codigo'), nullable=False)
    texto         = db.Column(db.String(500), nullable=False)
    ordem         = db.Column(db.Integer, default=0)
    concluido     = db.Column(db.Boolean, default=False)
    observacao    = db.Column(db.Text, nullable=True)
    concluido_por = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)
    concluido_em  = db.Column(db.DateTime, nullable=True)
    criado_por    = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)
    criado_em     = db.Column(db.DateTime, default=agora_br)
    autor         = db.relationship('Usuario', foreign_keys=[criado_por])
    executor      = db.relationship('Usuario', foreign_keys=[concluido_por])

    def to_dict(self):
        return {
            'id':                 self.id,
            'texto':              self.texto,
            'ordem':              self.ordem,
            'concluido':          self.concluido,
            'observacao':         self.observacao or '',
            'concluido_por_nome': self.executor.nome if self.executor else None,
            'concluido_em':       self.concluido_em.strftime('%d/%m/%Y %H:%M') if self.concluido_em else None,
            'criado_por_nome':    self.autor.nome if self.autor else None,
        }


class LogAcesso(db.Model):
    __tablename__ = 'log_acessos'
    id         = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    data_hora  = db.Column(db.DateTime, default=agora_br)
    usuario    = db.relationship('Usuario', backref='acessos')

    def to_dict(self):
        return {'data_hora': self.data_hora.strftime('%d/%m/%Y %H:%M:%S')}


# ─────────────────────────────────────────
# CHANGELOG
# ─────────────────────────────────────────
CATEGORIAS_CHANGELOG = ['feature', 'fix', 'improvement', 'removed']

class ChangelogEntry(db.Model):
    __tablename__ = 'changelog_entries'
    id           = db.Column(db.Integer, primary_key=True)
    categoria    = db.Column(db.String(20), nullable=False)   # feature | fix | improvement | removed
    titulo       = db.Column(db.String(200), nullable=False)
    descricao    = db.Column(db.Text, nullable=True)
    criado_em    = db.Column(db.DateTime, default=agora_br)
    criado_por   = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)
    autor        = db.relationship('Usuario', foreign_keys=[criado_por])

    def to_dict(self):
        return {
            'id':         self.id,
            'categoria':  self.categoria,
            'titulo':     self.titulo,
            'descricao':  self.descricao or '',
            'criado_em':  self.criado_em.strftime('%d/%m/%Y %H:%M'),
            'autor_nome': self.autor.nome if self.autor else 'Sistema'
        }


class Ticket(db.Model):
    __tablename__ = 'tickets'
    id          = db.Column(db.Integer, primary_key=True)
    tipo        = db.Column(db.String(20), nullable=False)    # erro | sugestao | outro
    descricao   = db.Column(db.Text, nullable=False)
    status      = db.Column(db.String(20), default='aberto')  # aberto | em_analise | resolvido
    resposta    = db.Column(db.Text, nullable=True)
    empresa     = db.Column(db.String(100), nullable=True)
    criado_por  = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    criado_em   = db.Column(db.DateTime, default=agora_br)
    resolvido_em= db.Column(db.DateTime, nullable=True)
    autor       = db.relationship('Usuario', foreign_keys=[criado_por])

    def to_dict(self):
        return {
            'id':           self.id,
            'tipo':         self.tipo,
            'descricao':    self.descricao,
            'status':       self.status,
            'resposta':     self.resposta or '',
            'empresa':      self.empresa or '',
            'criado_por':   self.criado_por,
            'autor_nome':   self.autor.nome if self.autor else '?',
            'criado_em':    self.criado_em.strftime('%d/%m/%Y %H:%M'),
            'resolvido_em': self.resolvido_em.strftime('%d/%m/%Y %H:%M') if self.resolvido_em else None,
        }


# ─────────────────────────────────────────
# EMAIL HELPERS
# ─────────────────────────────────────────
def _enviar_async(destinatarios, assunto, corpo_html):
    api_key = os.environ.get('BREVO_API_KEY')
    if not api_key:
        return
    body = {
        'sender':      {'name': 'TaskFlow', 'email': os.environ.get('EMAIL_REMETENTE', 'noreply@taskflow.app')},
        'to':          [{'email': e} for e in destinatarios],
        'subject':     assunto,
        'htmlContent': corpo_html
    }
    email_cc = os.environ.get('EMAIL_CC', '')
    if email_cc:
        body['cc'] = [{'email': e.strip()} for e in email_cc.split(',') if e.strip()]
    payload = _json.dumps(body).encode('utf-8')
    req = urllib.request.Request(
        'https://api.brevo.com/v3/smtp/email',
        data=payload,
        headers={'api-key': api_key, 'Content-Type': 'application/json'},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            print(f'[EMAIL] Enviado: {resp.status}')
    except Exception as e:
        print(f'[EMAIL] Erro ao enviar: {e}')


def enviar_email(destinatarios, assunto, corpo_html):
    if not os.environ.get('BREVO_API_KEY') or not destinatarios:
        return
    t = threading.Thread(target=_enviar_async, args=(destinatarios, assunto, corpo_html))
    t.daemon = True
    t.start()


def _template_base(titulo, conteudo):
    return f"""
    <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #f4f4f4; padding: 20px;">
        <div style="background: #0f172a; padding: 24px; border-radius: 8px 8px 0 0; text-align: center;">
            <h1 style="color: #ffffff; margin: 0; font-size: 22px;">TaskFlow</h1>
        </div>
        <div style="background: #ffffff; padding: 32px; border-radius: 0 0 8px 8px;">
            <h2 style="color: #0f172a; margin-top: 0;">{titulo}</h2>
            {conteudo}
            <div style="text-align: center; margin: 28px 0 8px;">
                <a href="{APP_URL}" style="display:inline-block;background:#3b82f6;color:#ffffff;text-decoration:none;padding:12px 28px;border-radius:8px;font-weight:bold;font-size:15px;">
                    👉 Clique aqui e veja sua tarefa
                </a>
            </div>
            <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 24px 0;">
            <p style="color: #94a3b8; font-size: 12px; margin: 0; text-align:center;">Email automatico do TaskFlow. Nao responda.</p>
        </div>
    </div>"""


def _emails_envolvidos_tarefa(tarefa, excluir_id=None):
    emails = set()
    for u in tarefa.responsaveis:
        if u.id != excluir_id:
            emails.add(u.email)
    for u in tarefa.admins_colabs:
        if u.id != excluir_id:
            emails.add(u.email)
    if tarefa.criado_por and tarefa.criado_por != excluir_id:
        criador = db.session.get(Usuario, tarefa.criado_por)
        if criador:
            emails.add(criador.email)
    return emails


def email_tarefa_criada(tarefa, responsaveis, admin_nome, admin_email):
    emails = set(u.email for u in responsaveis)
    if tarefa.empresa:
        masters = Usuario.query.filter_by(tipo_perfil='Admin Master', empresa=tarefa.empresa).all()
        for a in masters:
            if a.email != admin_email:
                emails.add(a.email)
    if not emails:
        return
    prioridade_cor = {'Alta': '#ef4444'}.get(tarefa.prioridade, '#64748b')
    resp_nomes = ', '.join(u.nome for u in responsaveis) if responsaveis else 'Nenhum'
    conteudo = f"""
        <p>Uma nova tarefa foi criada:</p>
        <div style="background:#f8fafc;border-left:4px solid #3b82f6;padding:16px;border-radius:4px;margin:16px 0">
            <p style="margin:0 0 8px"><strong>Tarefa:</strong> {tarefa.descricao}</p>
            <p style="margin:0 0 8px"><strong>Prioridade:</strong> <span style="color:{prioridade_cor};font-weight:bold">{tarefa.prioridade}</span></p>
            <p style="margin:0 0 8px"><strong>Responsaveis:</strong> {resp_nomes}</p>
            <p style="margin:0"><strong>Criada por:</strong> {admin_nome}</p>
        </div>"""
    enviar_email(list(emails), '[TaskFlow] Nova tarefa criada', _template_base('Nova tarefa criada', conteudo))


def email_tarefa_atribuida(tarefa, responsaveis_novos, admin_nome):
    if not responsaveis_novos:
        return
    prioridade_cor = {'Alta': '#ef4444'}.get(tarefa.prioridade, '#64748b')
    conteudo = f"""
        <p>Voce foi atribuido(a) a seguinte tarefa:</p>
        <div style="background:#f8fafc;border-left:4px solid #3b82f6;padding:16px;border-radius:4px;margin:16px 0">
            <p style="margin:0 0 8px"><strong>Tarefa:</strong> {tarefa.descricao}</p>
            <p style="margin:0 0 8px"><strong>Prioridade:</strong> <span style="color:{prioridade_cor};font-weight:bold">{tarefa.prioridade}</span></p>
            <p style="margin:0"><strong>Atribuida por:</strong> {admin_nome}</p>
        </div>"""
    enviar_email([u.email for u in responsaveis_novos], '[TaskFlow] Tarefa atribuida a voce', _template_base('Nova tarefa atribuida', conteudo))


def email_comentario_adicionado(tarefa, comentario_texto, autor_nome, autor_id):
    envolvidos = _emails_envolvidos_tarefa(tarefa, excluir_id=autor_id)
    if not envolvidos:
        return
    conteudo = f"""
        <p>Novo comentario na tarefa que voce acompanha:</p>
        <div style="background:#f8fafc;border-left:4px solid #3b82f6;padding:16px;border-radius:4px;margin:16px 0">
            <p style="margin:0"><strong>Tarefa:</strong> {tarefa.descricao}</p>
        </div>
        <div style="background:#f1f5f9;border-radius:4px;padding:16px;margin:16px 0">
            <p style="margin:0 0 4px;color:#64748b;font-size:12px"><strong>{autor_nome}</strong> comentou:</p>
            <p style="margin:0;color:#0f172a">{comentario_texto}</p>
        </div>"""
    enviar_email(list(envolvidos), '[TaskFlow] Novo comentario na tarefa', _template_base('Novo comentario', conteudo))


def email_status_alterado(tarefa, status_anterior, status_novo, alterado_por_nome, alterado_por_id):
    emails = list(_emails_envolvidos_tarefa(tarefa, excluir_id=alterado_por_id))
    if not emails:
        return
    STATUS_COR = {
        'Não iniciado': '#94a3b8', 'Iniciado': '#3b82f6', 'Em andamento': '#8b5cf6',
        'Pausado': '#f59e0b', 'Aguardo retorno': '#f97316', 'Finalizado': '#22c55e',
    }
    cor = STATUS_COR.get(status_novo, '#64748b')
    conteudo = f"""
        <p>O status de uma tarefa foi atualizado:</p>
        <div style="background:#f8fafc;border-left:4px solid {cor};padding:16px;border-radius:4px;margin:16px 0">
            <p style="margin:0 0 8px"><strong>Tarefa:</strong> {tarefa.descricao}</p>
            <p style="margin:0 0 8px"><strong>Antes:</strong> <span style="color:#64748b">{status_anterior}</span></p>
            <p style="margin:0 0 8px"><strong>Agora:</strong> <span style="color:{cor};font-weight:bold">{status_novo}</span></p>
            <p style="margin:0"><strong>Alterado por:</strong> {alterado_por_nome}</p>
        </div>"""
    enviar_email(emails, f'[TaskFlow] Status atualizado: "{status_novo}"', _template_base('Status da tarefa atualizado', conteudo))


# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────
def usuario_atual():
    """Retorna o usuário logado na sessão."""
    return db.session.get(Usuario, session['usuario_id'])


def verificar_empresa(admin, alvo):
    """Retorna True se o admin NÃO tem permissão sobre o alvo (empresa diferente)."""
    return bool(admin.empresa and alvo.empresa != admin.empresa)


def safe_commit():
    """Commit com rollback automático em caso de erro."""
    try:
        db.session.commit()
        return True
    except Exception as e:
        db.session.rollback()
        print(f'[DB] Erro no commit: {e}')
        return False


def registrar_historico(id_tarefa, id_usuario, texto):
    db.session.add(Comentario(id_tarefa=id_tarefa, id_usuario=id_usuario, texto=texto, tipo='historico'))


def registrar_acesso(uid):
    db.session.add(LogAcesso(usuario_id=uid))


# ─────────────────────────────────────────
# DECORADORES
# ─────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            return jsonify({'erro': 'Nao autenticado'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            return jsonify({'erro': 'Nao autenticado'}), 401
        u = usuario_atual()
        if not u or not u.is_admin():
            return jsonify({'erro': 'Acesso negado. Apenas administradores.'}), 403
        return f(*args, **kwargs)
    return decorated


def master_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            return jsonify({'erro': 'Nao autenticado'}), 401
        u = usuario_atual()
        if not u or not u.is_master():
            return jsonify({'erro': 'Acesso negado. Apenas Admin Master.'}), 403
        return f(*args, **kwargs)
    return decorated


# ─────────────────────────────────────────
# ROTA PRINCIPAL
# ─────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


# ─────────────────────────────────────────
# AUTENTICACAO
# ─────────────────────────────────────────
@app.route('/api/login', methods=['POST'])
def login():
    dados = request.json
    email = dados.get('email', '').strip().lower()
    senha = dados.get('senha', '')
    if not email or not senha:
        return jsonify({'erro': 'E-mail e senha sao obrigatorios'}), 400
    u = Usuario.query.filter_by(email=email).first()
    if not u or not u.verificar_senha(senha):
        return jsonify({'erro': 'E-mail ou senha incorretos'}), 401
    session['usuario_id'] = u.id
    registrar_acesso(u.id)
    safe_commit()
    return jsonify(u.to_dict()), 200


@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'mensagem': 'Logout realizado'}), 200


@app.route('/api/me', methods=['GET'])
@login_required
def me():
    u = usuario_atual()
    u.last_seen = agora_br()
    registrar_acesso(u.id)
    safe_commit()
    return jsonify(u.to_dict()), 200

@app.route('/api/ping', methods=['POST'])
@login_required
def ping():
    u = usuario_atual()
    u.last_seen = agora_br()
    safe_commit()
    return jsonify({'ok': True})


# ─────────────────────────────────────────
# TROCAR / REDEFINIR SENHA
# ─────────────────────────────────────────
@app.route('/api/trocar-senha', methods=['POST'])
@login_required
def trocar_senha():
    dados       = request.json
    senha_atual = dados.get('senha_atual', '')
    senha_nova  = dados.get('senha_nova', '')
    senha_conf  = dados.get('senha_confirmacao', '')
    if not senha_atual or not senha_nova or not senha_conf:
        return jsonify({'erro': 'Preencha todos os campos'}), 400
    valida, erro = validar_senha_forte(senha_nova)
    if not valida:
        return jsonify({'erro': erro}), 400
    if senha_nova != senha_conf:
        return jsonify({'erro': 'A confirmacao de senha nao confere'}), 400
    u = usuario_atual()
    if not u.verificar_senha(senha_atual):
        return jsonify({'erro': 'Senha atual incorreta'}), 401
    u.definir_senha(senha_nova)
    u.trocar_senha = False
    safe_commit()
    return jsonify({'mensagem': 'Senha alterada com sucesso!'}), 200


@app.route('/api/trocar-email', methods=['POST'])
@login_required
def trocar_email():
    dados      = request.json
    senha      = dados.get('senha', '').strip()
    email_novo = dados.get('email_novo', '').strip().lower()
    if not senha or not email_novo:
        return jsonify({'erro': 'Preencha todos os campos'}), 400
    if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email_novo):
        return jsonify({'erro': 'E-mail inválido'}), 400
    u = usuario_atual()
    if not u.verificar_senha(senha):
        return jsonify({'erro': 'Senha incorreta'}), 401
    if Usuario.query.filter(Usuario.email == email_novo, Usuario.id != u.id).first():
        return jsonify({'erro': 'Este e-mail já está em uso'}), 409
    u.email = email_novo
    session['usuario_email'] = email_novo
    safe_commit()
    return jsonify({'mensagem': 'E-mail alterado com sucesso!'}), 200


@app.route('/api/usuarios/<int:uid>/redefinir-senha', methods=['POST'])
@admin_required
def redefinir_senha(uid):
    admin   = usuario_atual()
    usuario = db.session.get(Usuario, uid)
    if not usuario:
        return jsonify({'erro': 'Usuario nao encontrado'}), 404
    if verificar_empresa(admin, usuario):
        return jsonify({'erro': 'Acesso negado'}), 403
    nova_senha = request.json.get('senha_nova', '')
    valida, erro = validar_senha_forte(nova_senha)
    if not valida:
        return jsonify({'erro': erro}), 400
    usuario.definir_senha(nova_senha)
    usuario.trocar_senha = True
    safe_commit()
    return jsonify({'mensagem': f'Senha de {usuario.nome} redefinida.'}), 200


# ─────────────────────────────────────────
# USUARIOS
# ─────────────────────────────────────────
@app.route('/api/usuarios', methods=['GET'])
@login_required
def listar_usuarios():
    u = usuario_atual()
    q = Usuario.query
    if u.empresa:
        q = q.filter_by(empresa=u.empresa)
    usuarios = q.all()
    # Conta tarefas ativas (não deletadas) por usuário
    from sqlalchemy import func
    codigos_resp = db.session.query(
        tarefa_responsaveis.c.usuario_id,
        func.count(tarefa_responsaveis.c.tarefa_codigo).label('total')
    ).join(Tarefa, Tarefa.codigo == tarefa_responsaveis.c.tarefa_codigo)\
     .filter(Tarefa.deletado_em == None, Tarefa.status != 'Finalizado')\
     .group_by(tarefa_responsaveis.c.usuario_id).all()
    contagem = {row.usuario_id: row.total for row in codigos_resp}
    result = []
    for usr in usuarios:
        d = usr.to_dict()
        d['tarefas_ativas'] = contagem.get(usr.id, 0)
        result.append(d)
    return jsonify(result)


@app.route('/api/usuarios/colaborativos', methods=['GET'])
@login_required
def listar_colaborativos():
    u = usuario_atual()
    q = Usuario.query.filter_by(tipo_perfil='Colaborativo')
    if u.empresa:
        q = q.filter_by(empresa=u.empresa)
    return jsonify([x.to_dict() for x in q.all()])


@app.route('/api/usuarios/online', methods=['GET'])
@login_required
def usuarios_online():
    """Retorna IDs dos usuários ativos nos últimos 5 minutos."""
    limite = agora_br() - timedelta(minutes=5)
    u = usuario_atual()
    q = Usuario.query.filter(Usuario.last_seen >= limite)
    if u.empresa:
        q = q.filter_by(empresa=u.empresa)
    return jsonify([x.id for x in q.all()])


@app.route('/api/usuarios/admins', methods=['GET'])
@admin_required
def listar_admins():
    u = usuario_atual()
    q = Usuario.query.filter(
        Usuario.tipo_perfil.in_(['Administrador', 'Admin Master']),
        Usuario.id != u.id
    )
    if u.empresa:
        q = q.filter_by(empresa=u.empresa)
    return jsonify([x.to_dict() for x in q.all()])


@app.route('/api/usuarios', methods=['POST'])
@admin_required
def criar_usuario():
    dados = request.json
    for campo in ['nome', 'funcao', 'email', 'tipo_perfil', 'senha']:
        if not dados.get(campo):
            return jsonify({'erro': f'Campo "{campo}" obrigatorio'}), 400
    perfis_validos = ['Colaborativo', 'Administrador', 'Admin Master']
    if dados['tipo_perfil'] not in perfis_validos:
        return jsonify({'erro': 'tipo_perfil invalido'}), 400
    criador = usuario_atual()
    if dados['tipo_perfil'] == 'Admin Master' and not criador.is_master():
        return jsonify({'erro': 'Apenas Admin Master pode criar outros Admin Masters'}), 403
    if dados['tipo_perfil'] == 'Admin Master' and criador.email != ADMIN_MASTER_EMAIL:
        return jsonify({'erro': 'Apenas o Admin Master principal pode criar outros Admin Masters'}), 403
    valida, erro = validar_senha_forte(dados['senha'])
    if not valida:
        return jsonify({'erro': erro}), 400
    if Usuario.query.filter_by(email=dados['email'].lower()).first():
        return jsonify({'erro': 'E-mail ja cadastrado'}), 409
    empresa = criador.empresa or dados.get('empresa') or None
    novo = Usuario(
        nome=dados['nome'], funcao=dados['funcao'],
        email=dados['email'].lower(), tipo_perfil=dados['tipo_perfil'],
        trocar_senha=True, empresa=empresa,
        setor=dados.get('setor') or None
    )
    novo.definir_senha(dados['senha'])
    db.session.add(novo)
    safe_commit()
    return jsonify(novo.to_dict()), 201


@app.route('/api/usuarios/<int:uid>', methods=['PUT'])
@master_required
def editar_usuario(uid):
    admin   = usuario_atual()
    usuario = db.session.get(Usuario, uid)
    if not usuario:
        return jsonify({'erro': 'Usuario nao encontrado'}), 404
    if verificar_empresa(admin, usuario):
        return jsonify({'erro': 'Acesso negado'}), 403
    dados = request.json
    if 'nome'   in dados: usuario.nome   = dados['nome'].strip()
    if 'funcao' in dados: usuario.funcao = dados['funcao'].strip()
    if 'setor'  in dados: usuario.setor  = dados['setor'].strip() or None
    if 'empresa' in dados and not admin.empresa:
        usuario.empresa = dados['empresa'].strip() or None
    if 'tipo_perfil' in dados:
        novo_perfil = dados['tipo_perfil']
        if novo_perfil not in ['Colaborativo', 'Administrador', 'Admin Master']:
            return jsonify({'erro': 'tipo_perfil invalido'}), 400
        if novo_perfil == 'Admin Master' and admin.email != ADMIN_MASTER_EMAIL:
            return jsonify({'erro': 'Apenas o Admin Master principal pode promover outros a Admin Master'}), 403
        if usuario.email == ADMIN_MASTER_EMAIL and novo_perfil != 'Admin Master':
            return jsonify({'erro': 'Nao e possivel alterar o perfil do Admin Master principal'}), 403
        usuario.tipo_perfil = novo_perfil
    safe_commit()
    return jsonify(usuario.to_dict()), 200


@app.route('/api/usuarios/<int:uid>', methods=['DELETE'])
@admin_required
def excluir_usuario(uid):
    admin   = usuario_atual()
    usuario = db.session.get(Usuario, uid)
    if not usuario:
        return jsonify({'erro': 'Usuario nao encontrado'}), 404
    if usuario.id == session['usuario_id']:
        return jsonify({'erro': 'Voce nao pode excluir a si mesmo'}), 400
    if verificar_empresa(admin, usuario):
        return jsonify({'erro': 'Acesso negado'}), 403
    if usuario.is_master() and not admin.is_master():
        return jsonify({'erro': 'Apenas Admin Master pode excluir outro Admin Master'}), 403
    db.session.delete(usuario)
    safe_commit()
    return jsonify({'mensagem': 'Usuario excluido'}), 200


@app.route('/api/usuarios/<int:uid>/acessos', methods=['GET'])
@admin_required
def listar_acessos(uid):
    admin   = usuario_atual()
    usuario = db.session.get(Usuario, uid)
    if not usuario:
        return jsonify({'erro': 'Usuario nao encontrado'}), 404
    if verificar_empresa(admin, usuario):
        return jsonify({'erro': 'Acesso negado'}), 403

    # Filtro por período opcional: ?de=2024-01-01&ate=2024-01-31
    de_str  = request.args.get('de')
    ate_str = request.args.get('ate')
    q = LogAcesso.query.filter_by(usuario_id=uid)
    if de_str:
        try:
            q = q.filter(LogAcesso.data_hora >= datetime.strptime(de_str, '%Y-%m-%d'))
        except ValueError:
            pass
    if ate_str:
        try:
            ate = datetime.strptime(ate_str, '%Y-%m-%d') + timedelta(days=1)
            q = q.filter(LogAcesso.data_hora < ate)
        except ValueError:
            pass
    logs = q.order_by(LogAcesso.data_hora.desc()).limit(200).all()
    return jsonify([l.to_dict() for l in logs]), 200


@app.route('/api/setores', methods=['GET'])
@login_required
def listar_setores():
    u     = usuario_atual()
    query = db.session.query(Usuario.setor).filter(Usuario.setor.isnot(None))
    if u.empresa:
        query = query.filter(Usuario.empresa == u.empresa)
    setores = sorted(set(row[0] for row in query.all() if row[0]))
    return jsonify(setores)


# ─────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────
@app.route('/api/dashboard', methods=['GET'])
@login_required
def dashboard():
    u       = usuario_atual()
    empresa = u.empresa
    uid     = u.id

    # Filtros opcionais
    filtro_status   = request.args.get('status')       # ex: "Iniciado" ou "pendentes"
    filtro_uid_resp = request.args.get('usuario_id', type=int)
    filtro_periodo  = request.args.get('periodo', '14', type=str)

    # Query base sem deletados
    if u.is_master():
        q = Tarefa.query.filter(
            db.or_(Tarefa.compartilhada == True, Tarefa.criado_por == uid),
            Tarefa.deletado_em == None
        )
        if empresa:
            q = q.filter(Tarefa.empresa == empresa)
    elif u.is_admin():
        q = Tarefa.query.filter(
            db.or_(
                Tarefa.criado_por == uid,
                Tarefa.codigo.in_(db.session.query(tarefa_admins.c.tarefa_codigo).filter(tarefa_admins.c.usuario_id == uid)),
                Tarefa.codigo.in_(db.session.query(tarefa_responsaveis.c.tarefa_codigo).filter(tarefa_responsaveis.c.usuario_id == uid))
            ),
            Tarefa.deletado_em == None
        )
        if empresa:
            q = q.filter(Tarefa.empresa == empresa)
    else:
        q = Tarefa.query.join(tarefa_responsaveis, Tarefa.codigo == tarefa_responsaveis.c.tarefa_codigo)\
            .filter(tarefa_responsaveis.c.usuario_id == uid, Tarefa.deletado_em == None)
        if empresa:
            q = q.filter(Tarefa.empresa == empresa)

    todas_tarefas = q.all()

    # Função central: pessoa está envolvida na tarefa?
    def tem_pessoa(t, pid):
        return (
            any(r.id == pid for r in t.responsaveis) or
            any(a.id == pid for a in t.admins_colabs) or
            t.criado_por == pid
        )

    STATUSES_PENDENTES = [s for s in STATUSES_VALIDOS if s != 'Finalizado']

    tarefas_filtradas = todas_tarefas
    if filtro_uid_resp:
        tarefas_filtradas = [t for t in todas_tarefas if tem_pessoa(t, filtro_uid_resp)]
    if filtro_status == 'pendentes':
        tarefas_filtradas = [t for t in tarefas_filtradas if t.status in STATUSES_PENDENTES]
    elif filtro_status:
        tarefas_filtradas = [t for t in tarefas_filtradas if t.status == filtro_status]

    total = len(tarefas_filtradas)

    por_status = {}
    for s in STATUSES_VALIDOS:
        por_status[s] = sum(1 for t in tarefas_filtradas if t.status == s)

    alta_prioridade = sum(1 for t in tarefas_filtradas if t.prioridade == 'Alta' and t.status != 'Finalizado')

    # Tarefas recentes (sem filtro de status, mas com filtro de usuário)
    base_recentes = todas_tarefas if not filtro_uid_resp else [t for t in todas_tarefas if tem_pessoa(t, filtro_uid_resp)]
    recentes = sorted(base_recentes, key=lambda t: t.codigo, reverse=True)[:8]

    # Gráfico por período escolhido
    dias = int(filtro_periodo) if filtro_periodo in ('7','14','30','90') else 14
    hoje = agora_br().date()
    criacoes_por_dia = {}
    for i in range(dias - 1, -1, -1):
        d = (hoje - timedelta(days=i)).strftime('%d/%m')
        criacoes_por_dia[d] = 0
    base_graf = todas_tarefas if not filtro_uid_resp else [t for t in todas_tarefas if tem_pessoa(t, filtro_uid_resp)]
    for t in base_graf:
        d = t.data_criacao.strftime('%d/%m')
        if d in criacoes_por_dia:
            criacoes_por_dia[d] += 1

    # Lista de TODAS as pessoas com tarefas (responsáveis + criadores + admins_colabs) para o filtro
    pessoas_set = {}
    for t in todas_tarefas:
        # Responsáveis colaborativos
        for r in t.responsaveis:
            if r.id not in pessoas_set:
                pessoas_set[r.id] = r.nome
        # Admins colaboradores adicionados à tarefa
        for a in t.admins_colabs:
            if a.id not in pessoas_set:
                pessoas_set[a.id] = a.nome
        # Criador da tarefa (admin/master com tarefas pessoais)
        if t.criado_por:
            criador = db.session.get(Usuario, t.criado_por)
            if criador and criador.id not in pessoas_set:
                pessoas_set[criador.id] = criador.nome
    responsaveis_lista = sorted([{'id': k, 'nome': v} for k, v in pessoas_set.items()], key=lambda x: x['nome'])

    # Por usuário (ranking de pendências) — inclui responsáveis, admins_colabs E criadores de tarefas pessoais
    por_usuario = {}
    for t in todas_tarefas:
        if t.status == 'Finalizado':
            continue
        ids_resp      = [r.id for r in t.responsaveis]
        ids_adm_colab = [a.id for a in t.admins_colabs]

        if ids_resp:
            # Tarefa compartilhada com colaborativos — conta para cada responsável
            for r in t.responsaveis:
                if filtro_uid_resp and r.id != filtro_uid_resp:
                    continue
                if r.id not in por_usuario:
                    por_usuario[r.id] = {'nome': r.nome, 'total': 0, 'alta': 0, 'id': r.id}
                por_usuario[r.id]['total'] += 1
                if t.prioridade == 'Alta':
                    por_usuario[r.id]['alta'] += 1
        elif ids_adm_colab:
            # Tarefa delegada a admins colaboradores
            for a in t.admins_colabs:
                if filtro_uid_resp and a.id != filtro_uid_resp:
                    continue
                if a.id not in por_usuario:
                    por_usuario[a.id] = {'nome': a.nome, 'total': 0, 'alta': 0, 'id': a.id}
                por_usuario[a.id]['total'] += 1
                if t.prioridade == 'Alta':
                    por_usuario[a.id]['alta'] += 1
        elif t.criado_por:
            # Tarefa pessoal — conta para o criador
            if filtro_uid_resp and t.criado_por != filtro_uid_resp:
                continue
            criador = db.session.get(Usuario, t.criado_por)
            if criador:
                if criador.id not in por_usuario:
                    por_usuario[criador.id] = {'nome': criador.nome, 'total': 0, 'alta': 0, 'id': criador.id}
                por_usuario[criador.id]['total'] += 1
                if t.prioridade == 'Alta':
                    por_usuario[criador.id]['alta'] += 1
    ranking = sorted(por_usuario.values(), key=lambda x: x['total'], reverse=True)[:10]

    # Lista de tarefas para exibição:
    # — Sem filtro: 8 mais recentes
    # — Com qualquer filtro: todas as tarefas filtradas (já via tem_pessoa + status)
    if filtro_uid_resp or filtro_status:
        base_lista = sorted(tarefas_filtradas, key=lambda t: t.codigo, reverse=True)
    else:
        base_lista = sorted(todas_tarefas, key=lambda t: t.codigo, reverse=True)[:8]

    return jsonify({
        'total':             total,
        'por_status':        por_status,
        'alta_prioridade':   alta_prioridade,
        'finalizadas':       por_status.get('Finalizado', 0),
        'pendentes':         total - por_status.get('Finalizado', 0),
        'recentes':          [{'codigo': t.codigo, 'descricao': t.descricao, 'status': t.status, 'prioridade': t.prioridade,
                               'responsaveis': [r.nome for r in t.responsaveis],
                               'criado_por_nome': db.session.get(Usuario, t.criado_por).nome if t.criado_por else None}
                              for t in base_lista],
        'criacoes_por_dia':  [{'dia': k, 'total': v} for k, v in criacoes_por_dia.items()],
        'responsaveis_lista': responsaveis_lista,
        'ranking_pendencias': ranking,
        'filtros_ativos':    {'status': filtro_status, 'usuario_id': filtro_uid_resp, 'periodo': dias}
    })


# ─────────────────────────────────────────
# TAREFAS
# ─────────────────────────────────────────
@app.route('/api/tarefas', methods=['GET'])
@login_required
def listar_tarefas():
    u       = usuario_atual()
    empresa = u.empresa
    uid     = u.id

    if u.is_master():
        q = Tarefa.query.filter(
            db.or_(Tarefa.compartilhada == True, Tarefa.criado_por == uid),
            Tarefa.deletado_em == None
        )
        if empresa:
            q = q.filter(Tarefa.empresa == empresa)
    elif u.is_admin():
        q = Tarefa.query.filter(
            db.or_(
                Tarefa.criado_por == uid,
                Tarefa.codigo.in_(db.session.query(tarefa_admins.c.tarefa_codigo).filter(tarefa_admins.c.usuario_id == uid)),
                Tarefa.codigo.in_(db.session.query(tarefa_responsaveis.c.tarefa_codigo).filter(tarefa_responsaveis.c.usuario_id == uid))
            ),
            Tarefa.deletado_em == None
        )
        if empresa:
            q = q.filter(Tarefa.empresa == empresa)
    else:
        q = Tarefa.query.join(tarefa_responsaveis, Tarefa.codigo == tarefa_responsaveis.c.tarefa_codigo)\
            .filter(tarefa_responsaveis.c.usuario_id == uid, Tarefa.deletado_em == None)
        if empresa:
            q = q.filter(Tarefa.empresa == empresa)

    tarefas = q.order_by(Tarefa.codigo.desc()).all()
    return jsonify([t.to_dict(viewer_id=uid) for t in tarefas])


@app.route('/api/tarefas/lixeira', methods=['GET'])
@admin_required
def listar_lixeira():
    admin   = usuario_atual()
    empresa = admin.empresa
    if admin.is_master():
        q = Tarefa.query.filter(Tarefa.deletado_em != None)
        if empresa:
            q = q.filter(Tarefa.empresa == empresa)
    else:
        uid = admin.id
        q = Tarefa.query.filter(Tarefa.criado_por == uid, Tarefa.deletado_em != None)
    tarefas = q.order_by(Tarefa.deletado_em.desc()).all()
    result = []
    for t in tarefas:
        d = t.to_dict(viewer_id=admin.id)
        if t.deletado_por:
            u = db.session.get(Usuario, t.deletado_por)
            d['deletado_por_nome'] = u.nome if u else 'Desconhecido'
        else:
            d['deletado_por_nome'] = None
        result.append(d)
    return jsonify(result)


@app.route('/api/tarefas', methods=['POST'])
@admin_required
def criar_tarefa():
    dados = request.json
    if not dados.get('descricao'):
        return jsonify({'erro': 'Descricao obrigatoria'}), 400
    prioridade = dados.get('prioridade', 'Nenhuma')
    if prioridade not in PRIORIDADES_VALIDAS:
        return jsonify({'erro': 'Prioridade invalida'}), 400
    admin   = usuario_atual()
    empresa = admin.empresa
    uids_resp = dados.get('responsaveis_ids', [])
    uids_adm  = dados.get('admins_ids', [])
    compartilhada = len(uids_resp) > 0
    nova = Tarefa(
        descricao=dados['descricao'], prioridade=prioridade,
        compartilhada=compartilhada, criado_por=session['usuario_id'], empresa=empresa
    )
    if dados.get('data_prazo'):
        try:
            from datetime import date
            nova.data_prazo = date.fromisoformat(dados['data_prazo'])
        except Exception:
            pass
    nova.recorrente = dados.get('recorrente') or None
    db.session.add(nova)
    db.session.flush()
    responsaveis_novos = []
    nomes = []
    for uid in uids_resp:
        u = db.session.get(Usuario, uid)
        if u and u.tipo_perfil == 'Colaborativo' and (not empresa or u.empresa == empresa):
            nova.responsaveis.append(u)
            responsaveis_novos.append(u)
            nomes.append(u.nome)
    for uid in uids_adm:
        u = db.session.get(Usuario, uid)
        if u and u.is_admin() and u.id != admin.id and (not empresa or u.empresa == empresa):
            nova.admins_colabs.append(u)
            if not nova.compartilhada:
                nova.compartilhada = True
    msg = f'Tarefa criada por {admin.nome}.'
    msg += f' Responsaveis: {", ".join(nomes)}.' if nomes else ' Tarefa pessoal.'
    registrar_historico(nova.codigo, session['usuario_id'], msg)
    safe_commit()
    email_tarefa_criada(nova, responsaveis_novos, admin.nome, admin.email)
    return jsonify(nova.to_dict()), 201


@app.route('/api/tarefas/<int:codigo>', methods=['DELETE'])
@admin_required
def excluir_tarefa(codigo):
    tarefa = db.session.get(Tarefa, codigo)
    admin  = usuario_atual()
    if not tarefa or tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa nao encontrada'}), 404
    if admin.empresa and tarefa.empresa != admin.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    # Soft delete
    tarefa.deletado_em  = agora_br()
    tarefa.deletado_por = session['usuario_id']
    registrar_historico(codigo, session['usuario_id'], f'Tarefa movida para lixeira por {admin.nome}.')
    safe_commit()
    return jsonify({'mensagem': f'Tarefa #{codigo} movida para lixeira'}), 200


@app.route('/api/tarefas/<int:codigo>/restaurar', methods=['POST'])
@admin_required
def restaurar_tarefa(codigo):
    tarefa = db.session.get(Tarefa, codigo)
    admin  = usuario_atual()
    if not tarefa or not tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa nao encontrada na lixeira'}), 404
    if admin.empresa and tarefa.empresa != admin.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    tarefa.deletado_em = None
    registrar_historico(codigo, session['usuario_id'], f'Tarefa restaurada da lixeira por {admin.nome}.')
    safe_commit()
    return jsonify({'mensagem': f'Tarefa #{codigo} restaurada'}), 200


@app.route('/api/tarefas/<int:codigo>/permanente', methods=['DELETE'])
@admin_required
def excluir_permanente(codigo):
    tarefa = db.session.get(Tarefa, codigo)
    admin  = usuario_atual()
    if not tarefa:
        return jsonify({'erro': 'Tarefa nao encontrada'}), 404
    if admin.empresa and tarefa.empresa != admin.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    db.session.delete(tarefa)
    safe_commit()
    return jsonify({'mensagem': f'Tarefa #{codigo} excluida permanentemente'}), 200


@app.route('/api/tarefas/<int:codigo>/status', methods=['PATCH'])
@login_required
def atualizar_status(codigo):
    tarefa  = db.session.get(Tarefa, codigo)
    usuario = usuario_atual()
    if not tarefa or tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa nao encontrada'}), 404
    if usuario.empresa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado — empresa diferente'}), 403
    # Colaborativo só altera status de tarefas onde é responsável
    if usuario.tipo_perfil == 'Colaborativo' and usuario.id not in [u.id for u in tarefa.responsaveis]:
        return jsonify({'erro': 'Acesso negado — você não é responsável desta tarefa'}), 403
    # Admin e Admin Master podem alterar status de qualquer tarefa da empresa
    novo_status = request.json.get('status')
    if novo_status not in STATUSES_VALIDOS:
        return jsonify({'erro': 'Status invalido'}), 400
    anterior = tarefa.status
    if novo_status == 'Não iniciado' and anterior != 'Não iniciado':
        registrar_historico(codigo, session['usuario_id'],
            f'{usuario.nome} tentou reverter o status para "Não iniciado" (bloqueado — tarefa já foi iniciada).')
        safe_commit()
        return jsonify({'erro': 'Não é possível reverter para "Não iniciado" após a tarefa ser iniciada.', 'bloqueado': True}), 400
    tarefa.status = novo_status
    registrar_historico(codigo, session['usuario_id'],
        f'Status alterado de "{anterior}" para "{novo_status}" por {usuario.nome}.')
    safe_commit()
    if anterior != novo_status:
        email_status_alterado(tarefa, anterior, novo_status, usuario.nome, usuario.id)

    # Tarefas recorrentes: ao finalizar, cria próxima ocorrência
    nova_tarefa = None
    if novo_status == 'Finalizado' and tarefa.recorrente in ('semanal', 'mensal'):
        from datetime import date, timedelta
        if tarefa.data_prazo:
            if tarefa.recorrente == 'semanal':
                proximo_prazo = tarefa.data_prazo + timedelta(weeks=1)
            else:
                # Avança um mês, mantendo o dia
                m = tarefa.data_prazo.month + 1
                y = tarefa.data_prazo.year + (m - 1) // 12
                m = ((m - 1) % 12) + 1
                import calendar
                d = min(tarefa.data_prazo.day, calendar.monthrange(y, m)[1])
                proximo_prazo = date(y, m, d)
        else:
            proximo_prazo = None

        nova_tarefa = Tarefa(
            descricao=tarefa.descricao,
            prioridade=tarefa.prioridade,
            compartilhada=tarefa.compartilhada,
            criado_por=tarefa.criado_por,
            empresa=tarefa.empresa,
            data_prazo=proximo_prazo,
            recorrente=tarefa.recorrente
        )
        db.session.add(nova_tarefa)
        db.session.flush()
        for r in tarefa.responsaveis:
            nova_tarefa.responsaveis.append(r)
        for a in tarefa.admins_colabs:
            nova_tarefa.admins_colabs.append(a)
        registrar_historico(nova_tarefa.codigo, session['usuario_id'],
            f'Tarefa recorrente ({tarefa.recorrente}) criada automaticamente a partir de #{codigo}.')
        safe_commit()

    resultado = tarefa.to_dict()
    if nova_tarefa:
        resultado['nova_recorrente'] = nova_tarefa.to_dict()
    return jsonify(resultado), 200


@app.route('/api/tarefas/<int:codigo>/prioridade', methods=['PATCH'])
@admin_required
def atualizar_prioridade(codigo):
    tarefa = db.session.get(Tarefa, codigo)
    admin  = usuario_atual()
    if not tarefa or tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa nao encontrada'}), 404
    if admin.empresa and tarefa.empresa != admin.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
        return jsonify({'erro': 'Acesso negado'}), 403
    nova = request.json.get('prioridade')
    if nova not in PRIORIDADES_VALIDAS:
        return jsonify({'erro': 'Prioridade invalida'}), 400
    tarefa.prioridade = nova
    safe_commit()
    return jsonify(tarefa.to_dict()), 200


@app.route('/api/tarefas/<int:codigo>/responsaveis', methods=['PUT'])
@admin_required
def atualizar_responsaveis(codigo):
    tarefa = db.session.get(Tarefa, codigo)
    admin  = usuario_atual()
    if not tarefa or tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa nao encontrada'}), 404
    if admin.empresa and tarefa.empresa != admin.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
        return jsonify({'erro': 'Acesso negado'}), 403
    ids_antes = {u.id for u in tarefa.responsaveis}
    empresa   = admin.empresa
    tarefa.responsaveis.clear()
    nomes_depois = []
    responsaveis_novos = []
    for uid in request.json.get('responsaveis_ids', []):
        u = db.session.get(Usuario, uid)
        if u and u.tipo_perfil == 'Colaborativo' and (not empresa or u.empresa == empresa):
            tarefa.responsaveis.append(u)
            nomes_depois.append(u.nome)
            if u.id not in ids_antes:
                responsaveis_novos.append(u)
    ids_depois = {u.id for u in tarefa.responsaveis}
    if ids_antes != ids_depois:
        nomes_antes_str = ', '.join(
            db.session.get(Usuario, i).nome for i in ids_antes if db.session.get(Usuario, i)
        ) or 'nenhum'
        registrar_historico(codigo, session['usuario_id'],
            f'Responsaveis alterados por {admin.nome}. Antes: {nomes_antes_str}. Agora: {", ".join(nomes_depois) or "nenhum"}.')
    safe_commit()
    if responsaveis_novos:
        email_tarefa_atribuida(tarefa, responsaveis_novos, admin.nome)
    return jsonify(tarefa.to_dict()), 200


@app.route('/api/tarefas/<int:codigo>/admins', methods=['PUT'])
@admin_required
def atualizar_admins_colab(codigo):
    tarefa = db.session.get(Tarefa, codigo)
    admin  = usuario_atual()
    if not tarefa or tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa nao encontrada'}), 404
    if admin.empresa and tarefa.empresa != admin.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
        return jsonify({'erro': 'Acesso negado'}), 403
    empresa = admin.empresa
    ids_admins_antes = {u.id for u in tarefa.admins_colabs}
    tarefa.admins_colabs.clear()
    nomes = []
    for uid in request.json.get('admins_ids', []):
        u = db.session.get(Usuario, uid)
        if u and u.is_admin() and u.id != admin.id and (not empresa or u.empresa == empresa):
            tarefa.admins_colabs.append(u)
            nomes.append(u.nome)
    if nomes and not tarefa.compartilhada:
        tarefa.compartilhada = True
    elif not nomes and not tarefa.responsaveis:
        tarefa.compartilhada = False
    ids_admins_depois = {u.id for u in tarefa.admins_colabs}
    if ids_admins_antes != ids_admins_depois:
        registrar_historico(codigo, session['usuario_id'],
            f'Admins colaboradores atualizados por {admin.nome}: {", ".join(nomes) or "nenhum"}.')
    safe_commit()
    return jsonify(tarefa.to_dict()), 200


# ─────────────────────────────────────────
# COMENTARIOS
# ─────────────────────────────────────────
@app.route('/api/tarefas/<int:codigo>/comentarios', methods=['GET'])
@login_required
def listar_comentarios(codigo):
    if not db.session.get(Tarefa, codigo):
        return jsonify({'erro': 'Tarefa nao encontrada'}), 404
    comentarios = Comentario.query.filter_by(id_tarefa=codigo).order_by(Comentario.data_hora.asc()).all()
    return jsonify([c.to_dict() for c in comentarios])


@app.route('/api/tarefas/<int:codigo>/comentarios', methods=['POST'])
@login_required
def adicionar_comentario(codigo):
    tarefa  = db.session.get(Tarefa, codigo)
    usuario = usuario_atual()
    if not tarefa or tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa nao encontrada'}), 404
    if usuario.empresa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    # Colaborativo só comenta em tarefas onde é responsável
    if usuario.tipo_perfil == 'Colaborativo':
        ids_resp = [u.id for u in tarefa.responsaveis]
        if usuario.id not in ids_resp:
            return jsonify({'erro': 'Acesso negado'}), 403
    # Admin e Admin Master podem comentar em qualquer tarefa da empresa
    texto = request.json.get('texto', '').strip()
    if not texto:
        return jsonify({'erro': 'Comentario nao pode ser vazio'}), 400
    novo = Comentario(id_tarefa=codigo, id_usuario=session['usuario_id'], texto=texto, tipo='comentario')
    db.session.add(novo)
    safe_commit()
    email_comentario_adicionado(tarefa, texto, usuario.nome, usuario.id)
    return jsonify(novo.to_dict()), 201


# ─────────────────────────────────────────
# ANEXOS
# ─────────────────────────────────────────
def extensao_permitida(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in EXTENSOES_PERMITIDAS

def formatar_tamanho(b):
    if b < 1024: return f'{b} B'
    if b < 1024**2: return f'{b/1024:.1f} KB'
    return f'{b/1024**2:.1f} MB'

def supabase_upload(file_bytes, nome_arquivo, mime_type):
    if not SUPABASE_URL or not SUPABASE_KEY:
        return False, 'Supabase não configurado'
    url = f'{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{nome_arquivo}'
    req = urllib.request.Request(
        url, data=file_bytes,
        headers={'Authorization': f'Bearer {SUPABASE_KEY}', 'Content-Type': mime_type, 'x-upsert': 'true'},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=30):
            return True, f'{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{nome_arquivo}'
    except Exception as e:
        return False, str(e)

def supabase_delete(nome_arquivo):
    if not SUPABASE_URL or not SUPABASE_KEY:
        return
    req = urllib.request.Request(
        f'{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{nome_arquivo}',
        headers={'Authorization': f'Bearer {SUPABASE_KEY}'}, method='DELETE'
    )
    try:
        urllib.request.urlopen(req, timeout=10)
    except Exception as e:
        print(f'[STORAGE] Erro ao deletar {nome_arquivo}: {e}')


@app.route('/api/tarefas/<int:codigo>/anexos', methods=['GET'])
@login_required
def listar_anexos(codigo):
    tarefa  = db.session.get(Tarefa, codigo)
    usuario = usuario_atual()
    if not tarefa:
        return jsonify({'erro': 'Tarefa não encontrada'}), 404
    if usuario.empresa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    return jsonify([a.to_dict() for a in tarefa.anexos])


@app.route('/api/tarefas/<int:codigo>/anexos', methods=['POST'])
@login_required
def upload_anexo(codigo):
    tarefa  = db.session.get(Tarefa, codigo)
    usuario = usuario_atual()
    if not tarefa or tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa não encontrada'}), 404
    if usuario.empresa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
    arquivo = request.files['arquivo']
    if not arquivo.filename or not extensao_permitida(arquivo.filename):
        return jsonify({'erro': 'Tipo de arquivo não permitido'}), 400
    nome_original = arquivo.filename
    ext       = nome_original.rsplit('.', 1)[1].lower() if '.' in nome_original else ''
    nome_base = re.sub(r'[^\w\-]', '_', nome_original.rsplit('.', 1)[0])[:40]
    nome_uuid = f'{nome_base}_{uuid.uuid4().hex[:8]}.{ext}' if ext else f'{nome_base}_{uuid.uuid4().hex[:8]}'
    mime_type = mimetypes.guess_type(nome_original)[0] or 'application/octet-stream'
    file_bytes = arquivo.read()
    ok, resultado = supabase_upload(file_bytes, nome_uuid, mime_type)
    if not ok:
        return jsonify({'erro': f'Erro no upload: {resultado}'}), 500
    novo = Anexo(
        id_tarefa=codigo, id_usuario=session['usuario_id'],
        nome_original=nome_original, nome_arquivo=resultado,
        tamanho=len(file_bytes), mime_type=mime_type
    )
    db.session.add(novo)
    registrar_historico(codigo, session['usuario_id'],
        f'{usuario.nome} anexou o arquivo "{nome_original}" ({formatar_tamanho(len(file_bytes))}).')
    safe_commit()
    return jsonify(novo.to_dict()), 201


@app.route('/api/anexos/<int:aid>/download', methods=['GET'])
@login_required
def download_anexo(aid):
    anexo   = db.session.get(Anexo, aid)
    usuario = usuario_atual()
    if not anexo:
        return jsonify({'erro': 'Anexo não encontrado'}), 404
    tarefa = db.session.get(Tarefa, anexo.id_tarefa)
    if usuario.empresa and tarefa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    if anexo.nome_arquivo.startswith('http'):
        return redirect(anexo.nome_arquivo)
    return jsonify({'erro': 'Arquivo não disponível'}), 404


@app.route('/api/anexos/<int:aid>', methods=['DELETE'])
@login_required
def excluir_anexo(aid):
    anexo   = db.session.get(Anexo, aid)
    usuario = usuario_atual()
    if not anexo:
        return jsonify({'erro': 'Anexo não encontrado'}), 404
    tarefa = db.session.get(Tarefa, anexo.id_tarefa)
    if not (anexo.id_usuario == usuario.id or (tarefa and tarefa.criado_por == usuario.id) or usuario.is_master()):
        return jsonify({'erro': 'Acesso negado'}), 403
    if anexo.nome_arquivo.startswith('http'):
        supabase_delete(anexo.nome_arquivo.split(f'/{SUPABASE_BUCKET}/')[-1])
    registrar_historico(anexo.id_tarefa, session['usuario_id'],
        f'{usuario.nome} removeu o anexo "{anexo.nome_original}".')
    db.session.delete(anexo)
    safe_commit()
    return jsonify({'mensagem': 'Anexo excluído'}), 200


# ─────────────────────────────────────────
# RELATÓRIO DE PENDÊNCIAS
# ─────────────────────────────────────────
@app.route('/api/relatorio/pendencias', methods=['GET'])
@admin_required
def relatorio_pendencias():
    admin      = usuario_atual()
    empresa    = admin.empresa
    filtro_uid = request.args.get('usuario_id', type=int)

    q = Tarefa.query.filter(Tarefa.status != 'Finalizado', Tarefa.deletado_em == None)
    if empresa:
        q = q.filter(Tarefa.empresa == empresa)
    if not admin.is_master():
        uid = admin.id
        q = q.filter(db.or_(
            Tarefa.criado_por == uid,
            Tarefa.codigo.in_(db.session.query(tarefa_admins.c.tarefa_codigo).filter(tarefa_admins.c.usuario_id == uid)),
            Tarefa.codigo.in_(db.session.query(tarefa_responsaveis.c.tarefa_codigo).filter(tarefa_responsaveis.c.usuario_id == uid))
        ))

    tarefas_pendentes = q.order_by(Tarefa.codigo.desc()).all()
    por_usuario = {}

    for t in tarefas_pendentes:
        if not t.compartilhada:
            continue
        ids_resp = [u.id for u in t.responsaveis]
        alvos = [db.session.get(Usuario, t.criado_por)] if not ids_resp else list(t.responsaveis)
        for u in alvos:
            if not u:
                continue
            uid = u.id
            if uid not in por_usuario:
                por_usuario[uid] = {'usuario': u.to_dict(), 'tarefas': []}
            por_usuario[uid]['tarefas'].append({
                'codigo':         t.codigo,
                'descricao':      t.descricao,
                'status':         t.status,
                'prioridade':     t.prioridade,
                'data_criacao':   t.data_criacao.strftime('%d/%m/%Y'),
                'sem_responsavel': not ids_resp
            })

    resultado = sorted(por_usuario.values(), key=lambda x: x['usuario']['nome'])
    if filtro_uid:
        resultado = [r for r in resultado if r['usuario']['id'] == filtro_uid]
    for item in resultado:
        item['tarefas'] = sorted(item['tarefas'], key=lambda t: (0 if t['prioridade'] == 'Alta' else 1, t['codigo']))
        item['total'] = len(item['tarefas'])

    return jsonify({
        'gerado_em':    agora_br().strftime('%d/%m/%Y %H:%M'),
        'total_tarefas': len(tarefas_pendentes),
        'por_usuario':  resultado
    })


# ─────────────────────────────────────────
# EXPORTAR RELATÓRIO EM EXCEL
# ─────────────────────────────────────────
@app.route('/api/relatorio/excel', methods=['GET'])
@admin_required
def relatorio_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'erro': 'openpyxl nao instalado. Adicione ao requirements.txt'}), 500

    admin      = usuario_atual()
    empresa    = admin.empresa
    filtro_uid = request.args.get('usuario_id', type=int)

    q = Tarefa.query.filter(Tarefa.status != 'Finalizado', Tarefa.deletado_em == None)
    if empresa:
        q = q.filter(Tarefa.empresa == empresa)
    if not admin.is_master():
        uid = admin.id
        q = q.filter(db.or_(
            Tarefa.criado_por == uid,
            Tarefa.codigo.in_(db.session.query(tarefa_admins.c.tarefa_codigo).filter(tarefa_admins.c.usuario_id == uid)),
            Tarefa.codigo.in_(db.session.query(tarefa_responsaveis.c.tarefa_codigo).filter(tarefa_responsaveis.c.usuario_id == uid))
        ))

    tarefas_pendentes = q.order_by(Tarefa.codigo.desc()).all()
    por_usuario = {}
    for t in tarefas_pendentes:
        if not t.compartilhada:
            continue
        ids_resp = [u.id for u in t.responsaveis]
        alvos = [db.session.get(Usuario, t.criado_por)] if not ids_resp else list(t.responsaveis)
        for u in alvos:
            if not u:
                continue
            uid = u.id
            if uid not in por_usuario:
                por_usuario[uid] = {'usuario': u, 'tarefas': []}
            por_usuario[uid]['tarefas'].append(t)

    resultado = sorted(por_usuario.values(), key=lambda x: x['usuario'].nome)
    if filtro_uid:
        resultado = [r for r in resultado if r['usuario'].id == filtro_uid]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pendências'

    # Cores
    COR_HEADER  = 'FF0F172A'
    COR_PESSOA  = 'FF1E3A5F'
    COR_ALT     = 'FFF0F4FF'
    STATUS_COR  = {
        'Não iniciado': 'FF94a3b8', 'Iniciado': 'FF3b82f6', 'Em andamento': 'FF8b5cf6',
        'Pausado': 'FFf59e0b', 'Aguardo retorno': 'FFf97316', 'Finalizado': 'FF22c55e'
    }

    thin = Side(style='thin', color='FFD0D9E8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells('A1:F1')
    titulo_cell = ws['A1']
    titulo_cell.value = f'Relatório de Pendências — {agora_br().strftime("%d/%m/%Y %H:%M")}'
    titulo_cell.font  = Font(bold=True, color='FFFFFFFF', size=13)
    titulo_cell.fill  = PatternFill('solid', fgColor=COR_HEADER)
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    row = 2
    colunas = ['#', 'Descrição', 'Status', 'Prioridade', 'Criação', 'Responsável']
    larguras = [8, 50, 18, 12, 12, 24]
    for col, (cab, larg) in enumerate(zip(colunas, larguras), 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    for item in resultado:
        u = item['usuario']
        tarefas = sorted(item['tarefas'], key=lambda t: (0 if t.prioridade == 'Alta' else 1, t.codigo))

        # Linha de pessoa
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = f'  {u.nome}  ·  {u.funcao}{" / " + u.setor if u.setor else ""}  ({len(tarefas)} pendência{"s" if len(tarefas) != 1 else ""})'
        cell.font  = Font(bold=True, color='FFFFFFFF', size=11)
        cell.fill  = PatternFill('solid', fgColor=COR_PESSOA)
        cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 22
        row += 1

        # Cabeçalho da tabela
        for col, cab in enumerate(colunas, 1):
            c = ws.cell(row=row, column=col, value=cab)
            c.font      = Font(bold=True, color='FF0F172A', size=10)
            c.fill      = PatternFill('solid', fgColor='FFE2E8F0')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = border
        ws.row_dimensions[row].height = 18
        row += 1

        for i, t in enumerate(tarefas):
            resp_nomes = ', '.join(u2.nome for u2 in t.responsaveis) if t.responsaveis else '—'
            valores = [
                f'#{str(t.codigo).zfill(4)}',
                t.descricao,
                t.status,
                t.prioridade,
                t.data_criacao.strftime('%d/%m/%Y'),
                resp_nomes
            ]
            fill_cor = 'FFFFFFFF' if i % 2 == 0 else COR_ALT
            for col, val in enumerate(valores, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill      = PatternFill('solid', fgColor=fill_cor)
                c.alignment = Alignment(vertical='center', wrap_text=(col == 2))
                c.border    = border
                c.font      = Font(size=10)
                if col == 3:  # Status com cor
                    cor_st = STATUS_COR.get(t.status, 'FF64748b')
                    c.font = Font(size=10, color=cor_st, bold=True)
                if col == 4 and t.prioridade == 'Alta':
                    c.font = Font(size=10, color='FFef4444', bold=True)
            ws.row_dimensions[row].height = 16
            row += 1

        row += 1  # linha em branco entre pessoas

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    nome_arquivo = f'relatorio_pendencias_{agora_br().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=nome_arquivo)


# ─────────────────────────────────────────
# CHECKLIST
# ─────────────────────────────────────────
def _pode_marcar_checklist(usuario, tarefa):
    ids_resp      = [u.id for u in tarefa.responsaveis]
    ids_adm_colab = [u.id for u in tarefa.admins_colabs]
    if usuario.is_master():
        return True
    if usuario.tipo_perfil == 'Administrador':
        return tarefa.criado_por == usuario.id or usuario.id in ids_adm_colab
    return usuario.id in ids_resp

def _pode_editar_checklist(usuario, tarefa):
    if not usuario.is_admin():
        return False
    if usuario.is_master():
        return True
    ids_adm_colab = [u.id for u in tarefa.admins_colabs]
    return tarefa.criado_por == usuario.id or usuario.id in ids_adm_colab


@app.route('/api/tarefas/<int:codigo>/checklist', methods=['GET'])
@login_required
def listar_checklist(codigo):
    tarefa  = db.session.get(Tarefa, codigo)
    usuario = usuario_atual()
    if not tarefa:
        return jsonify({'erro': 'Tarefa não encontrada'}), 404
    if usuario.empresa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    return jsonify([i.to_dict() for i in tarefa.checklist])


@app.route('/api/tarefas/<int:codigo>/checklist', methods=['POST'])
@login_required
def adicionar_item_checklist(codigo):
    tarefa  = db.session.get(Tarefa, codigo)
    usuario = usuario_atual()
    if not tarefa or tarefa.deletado_em:
        return jsonify({'erro': 'Tarefa não encontrada'}), 404
    if usuario.empresa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    if not _pode_editar_checklist(usuario, tarefa):
        return jsonify({'erro': 'Acesso negado'}), 403
    texto = request.json.get('texto', '').strip()
    if not texto:
        return jsonify({'erro': 'Texto obrigatório'}), 400
    max_ordem = db.session.query(db.func.max(ChecklistItem.ordem)).filter_by(id_tarefa=codigo).scalar() or 0
    item = ChecklistItem(id_tarefa=codigo, texto=texto, ordem=max_ordem + 1, criado_por=session['usuario_id'])
    db.session.add(item)
    safe_commit()
    return jsonify(item.to_dict()), 201


@app.route('/api/checklist/<int:item_id>', methods=['DELETE'])
@login_required
def remover_item_checklist(item_id):
    item    = db.session.get(ChecklistItem, item_id)
    usuario = usuario_atual()
    if not item:
        return jsonify({'erro': 'Item não encontrado'}), 404
    tarefa = db.session.get(Tarefa, item.id_tarefa)
    if usuario.empresa and tarefa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    if not _pode_editar_checklist(usuario, tarefa):
        return jsonify({'erro': 'Acesso negado'}), 403
    db.session.delete(item)
    safe_commit()
    return jsonify({'mensagem': 'Item removido'}), 200


@app.route('/api/checklist/<int:item_id>/marcar', methods=['PATCH'])
@login_required
def marcar_item_checklist(item_id):
    item    = db.session.get(ChecklistItem, item_id)
    usuario = usuario_atual()
    if not item:
        return jsonify({'erro': 'Item não encontrado'}), 404
    tarefa = db.session.get(Tarefa, item.id_tarefa)
    if usuario.empresa and tarefa and tarefa.empresa != usuario.empresa:
        return jsonify({'erro': 'Acesso negado'}), 403
    if not _pode_marcar_checklist(usuario, tarefa):
        return jsonify({'erro': 'Acesso negado'}), 403
    dados      = request.json
    concluido  = dados.get('concluido', False)
    observacao = dados.get('observacao', '').strip()
    item.concluido     = concluido
    item.observacao    = observacao if concluido else None
    item.concluido_por = session['usuario_id'] if concluido else None
    item.concluido_em  = agora_br() if concluido else None
    safe_commit()
    return jsonify(item.to_dict()), 200


# ─────────────────────────────────────────
# CHANGELOG ROUTES
# ─────────────────────────────────────────

@app.route('/api/changelog', methods=['GET'])
@login_required
def listar_changelog():
    entradas = ChangelogEntry.query.order_by(ChangelogEntry.criado_em.desc()).all()
    return jsonify([e.to_dict() for e in entradas])


@app.route('/api/changelog', methods=['POST'])
@login_required
def criar_changelog():
    usuario = usuario_atual()
    if not usuario.is_admin():
        return jsonify({'erro': 'Apenas administradores podem adicionar entradas'}), 403

    dados     = request.json
    categoria = dados.get('categoria', '').strip()
    titulo    = dados.get('titulo', '').strip()
    descricao = dados.get('descricao', '').strip()

    if not categoria or not titulo:
        return jsonify({'erro': 'Categoria e título são obrigatórios'}), 400
    if categoria not in CATEGORIAS_CHANGELOG:
        return jsonify({'erro': 'Categoria inválida'}), 400

    entrada = ChangelogEntry(
        categoria=categoria,
        titulo=titulo,
        descricao=descricao or None,
        criado_por=session['usuario_id']
    )
    db.session.add(entrada)
    safe_commit()
    return jsonify(entrada.to_dict()), 201


@app.route('/api/changelog/<int:entry_id>', methods=['DELETE'])
@login_required
def excluir_changelog(entry_id):
    usuario = usuario_atual()
    if not usuario.is_admin():
        return jsonify({'erro': 'Apenas administradores podem excluir entradas'}), 403
    entrada = db.session.get(ChangelogEntry, entry_id)
    if not entrada:
        return jsonify({'erro': 'Entrada não encontrada'}), 404
    db.session.delete(entrada)
    safe_commit()
    return jsonify({'mensagem': 'Entrada removida'}), 200


# ─────────────────────────────────────────
# TICKETS (feedback / bugs / sugestões)
# ─────────────────────────────────────────
@app.route('/api/tickets', methods=['POST'])
@login_required
def criar_ticket():
    u    = usuario_atual()
    dados = request.json or {}
    tipo  = dados.get('tipo', '').strip()
    desc  = dados.get('descricao', '').strip()
    if tipo not in ('erro', 'sugestao', 'outro'):
        return jsonify({'erro': 'Tipo inválido'}), 400
    if not desc:
        return jsonify({'erro': 'Descrição obrigatória'}), 400
    t = Ticket(tipo=tipo, descricao=desc, empresa=u.empresa, criado_por=u.id)
    db.session.add(t)
    safe_commit()
    return jsonify(t.to_dict()), 201


@app.route('/api/tickets', methods=['GET'])
@login_required
def listar_tickets():
    u = usuario_atual()
    if not u.is_master():
        return jsonify({'erro': 'Acesso negado'}), 403
    tickets = Ticket.query.order_by(Ticket.criado_em.desc()).all()
    return jsonify([t.to_dict() for t in tickets])


@app.route('/api/tickets/<int:tid>', methods=['PATCH'])
@login_required
def atualizar_ticket(tid):
    u = usuario_atual()
    if not u.is_master():
        return jsonify({'erro': 'Acesso negado'}), 403
    ticket = db.session.get(Ticket, tid)
    if not ticket:
        return jsonify({'erro': 'Ticket não encontrado'}), 404
    dados   = request.json or {}
    if 'status' in dados and dados['status'] in ('aberto', 'em_analise', 'resolvido'):
        ticket.status = dados['status']
        if dados['status'] == 'resolvido' and not ticket.resolvido_em:
            ticket.resolvido_em = agora_br()
        elif dados['status'] != 'resolvido':
            ticket.resolvido_em = None
    if 'resposta' in dados:
        ticket.resposta = dados['resposta'].strip()
    safe_commit()
    return jsonify(ticket.to_dict())


@app.route('/api/tickets/<int:tid>', methods=['DELETE'])
@login_required
def excluir_ticket(tid):
    u = usuario_atual()
    if not u.is_master():
        return jsonify({'erro': 'Acesso negado'}), 403
    ticket = db.session.get(Ticket, tid)
    if not ticket:
        return jsonify({'erro': 'Ticket não encontrado'}), 404
    db.session.delete(ticket)
    safe_commit()
    return jsonify({'mensagem': 'Ticket excluído'})


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5000)